# -*- coding: utf-8 -*-
"""
MOD v6 §2 — Génère deux fichiers Excel de démonstration pour la soutenance PFE :

  1. seed_candidats_demo.xlsx
     8 candidats fictifs (notes L1/L2/L3 + bac, redoublements, sessions) avec
     score calculé par la formule GL/DS officielle : Score = M.G + B.N.R + B.S.P
     + colonnes Avis Membre 1 / Avis Membre 2.

  2. modele_import_verification_inscription.xlsx
     Tableau de suivi « Inscription en ligne » : Attestation Déposée,
     Pièces Conformes, Statut Inscription Finale, Observations.

Sortie : services/candidature_service/demo_files/

Usage : python generate_demo_excel.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_DIR = os.path.join(os.path.dirname(__file__), 'demo_files')
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', fgColor='1A3A6B')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
CENTER = Alignment(horizontal='center', vertical='center')
THIN = Side(style='thin', color='DDE3ED')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 22


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 3, 12), 46)


def _bnr(redoub):
    return 5 if redoub == 0 else 3 if redoub == 1 else 0


def _bsp(sessions):
    return 3 if sessions == 0 else 2 if sessions == 1 else 0


# (num, nom, prenom, cin, l1, l2, l3, bac, redoub, sessions, avis1, avis2, statut)
CANDIDATS = [
    ("CAND-2026-001", "Ben Ali", "Ahmed",   "10234567", 13.50, 14.00, 15.01, 15.20, 0, 0, "Favorable",   "Favorable",   "Présélectionné"),
    ("CAND-2026-002", "Jellali", "Ranim",   "10234568", 12.10, 11.80, 12.40, 13.00, 1, 1, "Favorable",   "Défavorable", "Présélectionné"),
    ("CAND-2026-003", "Gharbi",  "Marwen",  "10234569", 11.20, 12.00, 11.50, 12.10, 0, 1, "Favorable",   "Favorable",   "Présélectionné"),
    ("CAND-2026-004", "Amira",   "Fatima",  "10234570", 15.30, 16.10, 15.80, 16.40, 0, 0, "Favorable",   "Favorable",   "Présélectionné"),
    ("CAND-2026-005", "Trabelsi","Yassine", "10234571", 10.50, 11.00, 10.80, 11.20, 2, 1, "Défavorable", "Défavorable", "Rejeté"),
    ("CAND-2026-006", "Mejri",   "Salma",   "10234572", 14.00, 13.60, 14.20, 14.00, 0, 0, "Favorable",   "Favorable",   "Présélectionné"),
    ("CAND-2026-007", "Bouazizi","Karim",   "10234573", 12.80, 12.20, 13.00, 12.50, 1, 0, "Favorable",   "Défavorable", "En attente"),
    ("CAND-2026-008", "Hammami", "Nour",    "10234574", 13.20, 14.40, 13.90, 14.80, 0, 1, "Favorable",   "Favorable",   "Présélectionné"),
]


def build_candidats():
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidats démo"
    headers = ["N° Candidature", "Nom", "Prénom", "N° CIN",
               "Moy L1", "Moy L2", "Moy L3", "Moy Bac",
               "Redoublements", "Sessions Contrôle",
               "M.G", "B.N.R", "B.S.P", "Score (GL/DS)",
               "Avis Membre 1", "Avis Membre 2", "Statut Présélection"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for c in CANDIDATS:
        (num, nom, prenom, cin, l1, l2, l3, bac, redoub, sess, a1, a2, statut) = c
        mg = round((l1 + l2 + l3) / 3, 2)
        bnr = _bnr(redoub)
        bsp = _bsp(sess)
        score = round(mg + bnr + bsp, 2)
        ws.append([num, nom, prenom, cin, l1, l2, l3, bac, redoub, sess,
                   mg, bnr, bsp, score, a1, a2, statut])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = BORDER
            if cell.column_letter not in ('B', 'C', 'O', 'P', 'Q'):
                cell.alignment = CENTER
    _autosize(ws)

    # Note de formule
    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1,
            value="Formule GL/DS : Score = M.G + B.N.R + B.S.P  |  M.G=(L1+L2+L3)/3  |  "
                  "B.N.R=5/3/0 (0/1/≥2 redoub.)  |  B.S.P=3/2/0 (0/1/≥2 sessions)").font = Font(italic=True, color='6B7A90')

    path = os.path.join(OUT_DIR, 'seed_candidats_demo.xlsx')
    wb.save(path)
    return path


def build_verification():
    wb = Workbook()
    ws = wb.active
    ws.title = "Vérification inscription"
    headers = ["N° Candidature", "Nom", "Prénom", "Master",
               "Attestation Déposée", "Pièces Conformes",
               "Statut Inscription Finale", "Observations Responsable"]
    ws.append(headers)
    _style_header(ws, len(headers))

    rows = [
        ("CAND-2026-001", "Ben Ali", "Ahmed",  "Master Génie Logiciel", "Oui", "Oui", "Admis",       ""),
        ("CAND-2026-002", "Jellali", "Ranim",  "Master Génie Logiciel", "Oui", "Non", "En attente",  "Relevé de notes manquant"),
        ("CAND-2026-003", "Gharbi",  "Marwen", "Master Génie Logiciel", "Non", "Non", "En attente",  "Attestation non déposée"),
        ("CAND-2026-004", "Amira",   "Fatima", "Master Sciences des Données", "Oui", "Oui", "Admis",  ""),
    ]
    for r in rows:
        ws.append(list(r))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = BORDER
            if cell.column_letter in ('E', 'F', 'G'):
                cell.alignment = CENTER

    # Validation de données (listes déroulantes) pour la démo
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_oui_non = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    dv_statut = DataValidation(type="list", formula1='"Admis,En attente,Rejeté"', allow_blank=True)
    ws.add_data_validation(dv_oui_non)
    ws.add_data_validation(dv_statut)
    last = ws.max_row
    dv_oui_non.add(f"E2:F{last}")
    dv_statut.add(f"G2:G{last}")

    _autosize(ws)
    path = os.path.join(OUT_DIR, 'modele_import_verification_inscription.xlsx')
    wb.save(path)
    return path


if __name__ == '__main__':
    p1 = build_candidats()
    p2 = build_verification()
    print("OK")
    print(" -", p1)
    print(" -", p2)
